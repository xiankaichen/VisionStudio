# +----------------------------------------------------------------------
# | DjangoAdmin敏捷开发框架 [ 赋能开发者，助力企业发展 ]
# +----------------------------------------------------------------------
# | 版权所有 2021~2024 北京DjangoAdmin研发中心
# +----------------------------------------------------------------------
# | Licensed Apache-2.0 DjangoAdmin并不是自由软件，未经许可禁止去掉相关版权
# +----------------------------------------------------------------------
# | 官方网站: https://www.djangoadmin.cn
# +----------------------------------------------------------------------
# | 作者: @一米阳光 团队荣誉出品
# +----------------------------------------------------------------------
# | 版权和免责声明:
# | 本团队对该软件框架产品拥有知识产权（包括但不限于商标权、专利权、著作权、商业秘密等）
# | 均受到相关法律法规的保护，任何个人、组织和单位不得在未经本团队书面授权的情况下对所授权
# | 软件框架产品本身申请相关的知识产权，禁止用于任何违法、侵害他人合法权益等恶意的行为，禁
# | 止用于任何违反我国法律法规的一切项目研发，任何个人、组织和单位用于项目研发而产生的任何
# | 意外、疏忽、合约毁坏、诽谤、版权或知识产权侵犯及其造成的损失 (包括但不限于直接、间接、
# | 附带或衍生的损失等)，本团队不承担任何法律责任，本软件框架禁止任何单位和个人、组织用于
# | 任何违法、侵害他人合法利益等恶意的行为，如有发现违规、违法的犯罪行为，本团队将无条件配
# | 合公安机关调查取证同时保留一切以法律手段起诉的权利，本软件框架只能用于公司和个人内部的
# | 法律所允许的合法合规的软件产品研发，详细声明内容请阅读《框架免责声明》附件；
# +----------------------------------------------------------------------

import json
import logging
import os
import random
import time

from django.core.paginator import Paginator, PageNotAnInteger, InvalidPage, EmptyPage
from openpyxl.reader.excel import load_workbook
from openpyxl.workbook import Workbook

from application.level import forms
from application.level.models import Level
from application.upload import services
from config.env import DJANGO_TEMP_PATH, DJANGO_UPLOAD_DIR, DJANGO_IMAGE_URL
from constant.constants import PAGE_LIMIT
from utils import R, regular
from utils.file import mkdir

from utils.utils import uid


# 查询职级分页数据
def LevelList(request):
    # 页码
    page = int(request.GET.get('page', 1))
    # 每页数
    limit = int(request.GET.get('limit', PAGE_LIMIT))
    # 查询数据
    query = Level.objects.filter(is_delete=False)
    # 职级名称模糊筛选
    name = request.GET.get('name')
    if name:
        query = query.filter(name__contains=name)
    # 职级状态筛选
    status = request.GET.get('status')
    if status:
        query = query.filter(status=status)
    # 排序
    query = query.order_by("sort")
    # 分页设置
    paginator = Paginator(query, limit)
    # 记录总数
    count = paginator.count
    # 分页查询
    try:
        level_list = paginator.page(page)
        # todo: 注意捕获异常
    except PageNotAnInteger:
        # 如果请求的页数不是整数, 返回第一页。
        level_list = paginator.page(1)
    except InvalidPage:
        # 如果请求的页数不存在, 重定向页面
        return R.failed('找不到页面的内容')
    except EmptyPage:
        # 如果请求的页数不在合法的页数范围内，返回结果的最后一页。
        level_list = paginator.page(paginator.num_pages)
    # 遍历数据源
    result = []
    if len(level_list) > 0:
        for item in level_list:
            data = {
                'id': item.id,
                'name': item.name,
                'sort': item.sort,
                'status': item.status,
                'create_time': str(item.create_time.strftime('%Y-%m-%d %H:%M:%S')) if item.create_time else None,
                'update_time': str(item.update_time.strftime('%Y-%m-%d %H:%M:%S')) if item.update_time else None,
            }
            result.append(data)
    # 返回结果
    return R.ok(data=result, count=count)


# 根据ID查询职级详情
def LevelDetail(level_id):
    # 根据ID查询职级
    level = Level.objects.filter(is_delete=False, id=level_id).first()
    # 查询结果为空判断
    if not level:
        return None
    # 声明结构体
    data = {
        'id': level.id,
        'name': level.name,
        'status': level.status,
        'sort': level.sort,
    }
    # 返回结果
    return data


# 添加职级
def LevelAdd(request):
    try:
        # 接收请求参数
        json_data = request.body.decode()
        # 参数为空判断
        if not json_data:
            return R.failed("参数不能为空")
        # 数据类型转换
        dict_data = json.loads(json_data)
    except Exception as e:
        logging.info("错误信息：\n{}".format(e))
        return R.failed("参数错误")

    # 表单验证
    form = forms.LevelForm(data=dict_data)
    if form.is_valid():
        # 职级名称
        name = form.cleaned_data.get("name")
        # 职级状态
        status = int(dict_data.get("status"))
        # 职级排序
        sort = int(dict_data.get("sort"))
        # 创建数据
        Level.objects.create(
            name=name,
            status=status,
            sort=sort,
            create_user=uid(request)
        )
        # 返回结果
        return R.ok(msg="添加成功")
    else:
        # 获取错误描述
        err_msg = regular.get_err(form)
        # 返回错误信息
        return R.failed(msg=err_msg)


# 更新职级
def LevelUpdate(request):
    try:
        # 接收请求参数
        json_data = request.body.decode()
        # 参数判空
        if not json_data:
            return R.failed("参数不能为空")
        # 数据类型转换
        dict_data = json.loads(json_data)
        # 职级ID
        level_id = dict_data.get('id')
        # 职级ID判空
        if not level_id or int(level_id) <= 0:
            return R.failed("职级ID不能为空")
    except Exception as e:
        logging.info("错误信息：\n{}".format(e))
        return R.failed("参数错误")

    # 表单验证
    form = forms.LevelForm(data=dict_data)
    if form.is_valid():
        # 职级名称
        name = form.cleaned_data.get("name")
        # 职级状态
        status = int(form.cleaned_data.get("status"))
        # 职级排序
        sort = int(form.cleaned_data.get("sort"))
    else:
        # 获取错误描述
        err_msg = regular.get_err(form)
        # 返回错误信息
        return R.failed(msg=err_msg)

    # 根据ID查询职级
    level = Level.objects.only('id').filter(id=level_id, is_delete=False).first()
    # 查询结果判空
    if not level:
        return R.failed("记录不存在")
    # 给对象赋值
    level.name = name
    level.status = status
    level.sort = sort
    level.update_user = uid(request)
    # 更新记录
    level.save()
    # 返回结果
    return R.ok(msg="更新成功")


# 删除职级
def LevelDelete(level_id):
    # 记录ID为空判断
    if not level_id:
        return R.failed("记录ID不存在")
    # 分裂字符串
    list = level_id.split(',')
    # 计数器
    count = 0
    # 遍历数据源
    if len(list) > 0:
        for id in list:
            # 根据ID查询记录
            level = Level.objects.only('id').filter(id=int(id), is_delete=False).first()
            # 查询结果判空
            if not level:
                return R.failed("职级不存在")
            # 设置删除标识
            level.is_delete = True
            # 更新记录
            level.save()
            # 计数器+1
            count += 1
    # 返回结果
    return R.ok(msg="本次共删除{0}条数据".format(count))


# 设置状态
def LevelStatus(request):
    try:
        # 接收请求参数
        json_data = request.body.decode()
        # 参数判空
        if not json_data:
            return R.failed("参数不能为空")
        # 数据类型转换
        dict_data = json.loads(json_data)
    except Exception as e:
        logging.info("错误信息：\n{}".format(e))
        return R.failed("参数错误")

    # 表单验证
    form = forms.LevelStatusForm(data=dict_data)
    if form.is_valid():
        # 职级ID
        level_id = int(form.cleaned_data.get('id'))
        # 职级状态
        status = int(form.cleaned_data.get("status"))
    else:
        # 获取错误描述
        err_msg = regular.get_err(form)
        # 返回错误信息
        return R.failed(msg=err_msg)

    # 根据ID查询职级
    level = Level.objects.only('id').filter(id=level_id, is_delete=False).first()
    # 查询结果判空
    if not level:
        return R.failed("记录不存在")
    # 给对象赋值
    level.status = status
    # 更新记录
    level.save()
    # 返回结果
    return R.ok()


# 导入Excel
def LevelImport(request):
    # 读取上传的文件
    file = services.uploadFile(request, ".xls|.xlsx")
    # 本地文件地址
    path = file['filePath']
    # 打开工作簿
    wb = load_workbook(path)
    # 获取表单
    ws = wb.worksheets[0]
    # 读取最大行
    max_row = ws.max_row
    # 计数器
    count = 0
    # 遍历读取数据源
    for item in range(2, max_row + 1):
        # 职级名称
        name = ws.cell(item, 2).value
        # 职级排序
        sort = ws.cell(item, 3).value
        # 职级状态
        status = ws.cell(item, 4).value
        # 创建数据
        Level.objects.create(
            name=name,
            status=1 if status == "正常" else 2,
            sort=sort,
            create_user=uid(request)
        )
        count += 1
    # 返回结果
    return R.ok(msg="本次共导入【{}】条数据".format(count))


# 导出Excel
def LevelExport(request):
    # 查询数据
    query = Level.objects.filter(is_delete=False)
    # 职级名称模糊筛选
    name = request.GET.get('name')
    if name:
        query = query.filter(name__contains=name)
    # 职级状态筛选
    status = request.GET.get('status')
    if status:
        query = query.filter(status=status)
    # 排序
    query = query.order_by("sort")
    # 查询职级列表
    level_list = query.values()
    # 遍历数据源
    data_list = []
    if len(level_list) > 0:
        for item in level_list:
            data = {
                'id': item['id'],
                'name': item['name'],
                'sort': item['sort'],
                'status': "正常" if item['status'] == 1 else "禁用",
                'create_time': str(item['create_time'].strftime('%Y-%m-%d %H:%M:%S')) if item['create_time'] else None,
                'update_time': str(item['update_time'].strftime('%Y-%m-%d %H:%M:%S')) if item['update_time'] else None,
            }
            data_list.append(data)

    # 实例化工作薄
    wb = Workbook()
    # 指定第一个sheet
    ws = wb.worksheets[0]
    # 添加表头
    ws.append(['职级ID', '职级名称', '职级排序', '职级状态', '创建时间', '更新时间'])

    # 遍历外层列表
    for item in data_list:
        item_list = []
        # 遍历内层每一个字典dict，把dict每一个值存入list
        for item in item.values():
            item_list.append(item)
        # 加入数据
        ws.append(item_list)

    # 保存文件
    save_path = DJANGO_TEMP_PATH + "/" + time.strftime('%Y%m%d')
    # 创建存放目录
    mkdir(save_path)
    # 定义文件名，年月日时分秒随机数
    file_name = time.strftime('%Y%m%d%H%M%S') + '%05d' % random.randint(0, 100)
    # 重写合成文件名
    file_name = os.path.join(file_name + ".xlsx")
    # 拼接文件路径
    path = save_path + "/" + file_name
    # 写入文件
    wb.save(r'' + path + '')
    # 文件访问地址
    file_url = DJANGO_IMAGE_URL + path.replace(DJANGO_UPLOAD_DIR, "")
    # 返回结果
    return R.ok(data=file_url)


# 获取职级数据列表
def getLevelList():
    # 查询职级数据列表
    list = Level.objects.filter(is_delete=False, status=1).order_by("sort").values()
    # 实例化列表
    level_list = []
    # 遍历数据源
    if list:
        for v in list:
            # 加入数组
            level_list.append(v)
    # 返回结果
    return level_list
